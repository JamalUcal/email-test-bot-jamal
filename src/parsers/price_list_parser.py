"""
Price list parser for supplier attachments.

Parses Excel and CSV files using intelligent header detection.
"""

import pandas as pd
import csv
import json
import re
from typing import Dict, List, Optional, cast, Any, Literal, Tuple
from dataclasses import dataclass
from pathlib import Path

from utils.logger import get_logger
from utils.exceptions import ParsingError
from parsers.header_detector import HeaderDetector, DetectedHeaders
from parsers.field_name_detector import FieldNameDetector

logger = get_logger(__name__)


@dataclass
class PriceListItem:
    """Single item from a price list."""
    part_number: str
    description: str
    price: float
    former_part_number: Optional[str] = None
    supersede_part_number: Optional[str] = None
    brand: Optional[str] = None
    location: Optional[str] = None
    currency: Optional[str] = None
    row_number: Optional[int] = None


@dataclass
class ParsedPriceList:
    """Parsed price list with metadata."""
    items: List[PriceListItem]
    supplier: str
    brand: str
    location: str
    currency: str
    filename: str
    total_rows: int
    valid_rows: int
    errors: List[str]


class PriceListParser:
    """Parses price lists from Excel/CSV files."""
    
    # Chunk size for processing large files
    CHUNK_SIZE = 10000  # Process 10k rows at a time
    LOG_INTERVAL = 10000  # Log progress every 10k rows
    
    # Default price prefixes (fallback if no currency_config provided)
    # Include Rs./Rs so Indian price lists always have prefix stripped
    DEFAULT_PRICE_PREFIXES = ['$', '€', '£', 'Rs.', 'Rs']
    
    def __init__(
        self, 
        column_mapping_config: Dict[str, Any],
        currency_config: Optional[Dict[str, Any]] = None
    ):
        """
        Initialize parser with intelligent header and field detection.
        
        Args:
            column_mapping_config: Configuration for header/field detection (REQUIRED)
            currency_config: Currency configuration for price prefix stripping (optional)
        """
        self.header_detector = HeaderDetector(column_mapping_config)
        self.field_name_detector = FieldNameDetector(column_mapping_config)
        
        # Build price prefixes list from currency config
        self.price_prefixes = self._build_price_prefixes(currency_config)
        
        # Get parsing config for blank row handling
        parsing_config = column_mapping_config.get('parsing', {})
        self.max_consecutive_blank_rows = parsing_config.get('max_consecutive_blank_rows', 100)
        
        logger.info(
            "Initialized parser with intelligent header and field detection enabled",
            price_prefixes_count=len(self.price_prefixes),
            max_consecutive_blank_rows=self.max_consecutive_blank_rows
        )
    
    def _build_price_prefixes(self, currency_config: Optional[Dict[str, Any]]) -> List[str]:
        """
        Build list of price prefixes to strip from currency config.
        
        Collects all price_prefixes from each currency in the config.
        Falls back to default prefixes if no config provided.
        
        Args:
            currency_config: Currency configuration dict
            
        Returns:
            List of price prefixes to strip, sorted by length (longest first)
        """
        if not currency_config:
            logger.debug("No currency config provided, using default price prefixes")
            return self.DEFAULT_PRICE_PREFIXES.copy()
        
        prefixes: List[str] = []
        supported_currencies = currency_config.get('supported_currencies', {})
        
        for currency_code, currency_info in supported_currencies.items():
            # Add price_prefixes if defined
            currency_prefixes = currency_info.get('price_prefixes', [])
            for prefix in currency_prefixes:
                if prefix and prefix not in prefixes:
                    prefixes.append(prefix)
            
            # Also add the symbol as a fallback if no price_prefixes defined
            if not currency_prefixes:
                symbol = currency_info.get('symbol')
                if symbol and symbol not in prefixes:
                    prefixes.append(symbol)
        
        # Always include Indian Rupee prefixes so INR-style files are handled even if config is incomplete
        for prefix in ('Rs.', 'Rs'):
            if prefix not in prefixes:
                prefixes.append(prefix)
        
        # Sort by length (longest first) to avoid partial matches
        # e.g., strip "Rs." before "Rs" to avoid leaving trailing "."
        prefixes.sort(key=len, reverse=True)
        
        logger.debug(
            f"Built price prefixes from config",
            prefix_count=len(prefixes),
            prefixes=prefixes[:10]  # Log first 10
        )
        
        return prefixes if prefixes else self.DEFAULT_PRICE_PREFIXES.copy()
    
    def _extract_currency_from_format(self, number_format: str) -> Optional[str]:
        """
        Extract currency symbol or code from Excel number format string.
        
        Handles patterns like:
        - _("$"* #,##0.00_) → $
        - [$€-407]#,##0.00 → €
        - #,##0.00 "USD" → USD
        - $#,##0.00 → $
        
        Args:
            number_format: Excel number format string
            
        Returns:
            Extracted currency symbol/code or None
        """
        if not number_format or number_format == 'General':
            return None
        
        # Pattern 1: _("$"* format) - quoted symbols
        # Match: _("$"* or _("€"* or _("£"*
        quoted_match = re.search(r'_\("([^"]+)"\*', number_format)
        if quoted_match:
            return quoted_match.group(1)
        
        # Pattern 2: [$SYMBOL-locale] format
        # Match: [$€-407] or [$£-809]
        bracket_match = re.search(r'\[\$([^\-\]]+)', number_format)
        if bracket_match:
            return bracket_match.group(1)
        
        # Pattern 3: Quoted currency code
        # Match: "USD" or "EUR"
        code_match = re.search(r'"([A-Z]{3})"', number_format)
        if code_match:
            return code_match.group(1)
        
        # Pattern 4: Leading symbol (simple formats)
        # Match: $#,##0.00 or €#,##0.00
        if number_format and number_format[0] in ['$', '€', '£', '¥', '₹', '฿']:
            return number_format[0]
        
        return None
    
    def peek_file_for_currency(
        self,
        file_path: str,
        currency_detector: Any,
        allowed_currencies: Optional[List[str]] = None
    ) -> Tuple[Optional[str], Optional[DetectedHeaders]]:
        """
        Peek at file to detect currency from header or first data row.
        
        Lightweight inspection without full parsing. Checks:
        1. Header row for parameterized currency codes (e.g., "Price-USD")
        2. First data row price cell number format (e.g., _("$"* #,##0.00_)) [Excel only]
        3. First data row price value for currency symbols (e.g., "$100")
        
        Args:
            file_path: Path to file
            currency_detector: CurrencyDetector instance
            allowed_currencies: Optional list of allowed currency codes to restrict detection
            
        Returns:
            Tuple of (currency_code, detected_headers):
            - currency_code: Currency code if detected, None otherwise
            - detected_headers: DetectedHeaders object for reuse, None if detection failed
        """
        file_path_obj = Path(file_path)
        file_ext = file_path_obj.suffix.lower()
        
        logger.info(f"Peeking file for currency detection: {file_path_obj.name}")
        
        try:
            # Read header row and first data row
            if file_ext in ['.xlsx', '.xls', '.xlsb']:
                # Check if it's actually JSON
                if self._is_json_file(file_path):
                    logger.debug("File is JSON, currency detection not applicable")
                    return None, None
                
                # Determine the correct engine based on file extension
                if file_ext == '.xlsb':
                    engine = 'pyxlsb'
                elif file_ext == '.xls':
                    engine = 'xlrd'
                else:
                    engine = 'openpyxl'
                
                df_peek = pd.read_excel(
                    file_path,
                    nrows=self.header_detector.max_blank_rows * 3,  # Match header_detector read depth
                    header=None,
                    engine=engine
                )
            elif file_ext in ['.csv', '.txt']:
                df_peek = pd.read_csv(
                    file_path,
                    nrows=self.header_detector.max_blank_rows * 3,  # Match header_detector read depth
                    header=None,
                    encoding='utf-8',
                    sep=None,  # Auto-detect delimiter for .txt files
                    engine='python'  # Required for sep=None
                )
            else:
                logger.warning(f"Unsupported file type for currency peek: {file_ext}")
                return None, None
            
            # Try to detect headers with currency detector
            try:
                detected_headers = self.header_detector.detect_headers(
                    file_path,
                    matched_brand_text=None,
                    currency_detector=currency_detector,
                    allowed_currencies=allowed_currencies
                )
                
                # Check if currency was detected from header
                if detected_headers.matched_currency_code:
                    logger.info(
                        f"Currency detected from header: {detected_headers.matched_currency_code}",
                        file=file_path_obj.name
                    )
                    return detected_headers.matched_currency_code, detected_headers
                
                # Layer 4a: Check if a currency column was detected
                currency_col_index = detected_headers.column_indices.get('currency')
                if currency_col_index is not None:
                    # Read first data row to get currency value
                    data_row_index = detected_headers.header_row_index + 1
                    if data_row_index < len(df_peek):
                        first_data_row = df_peek.iloc[data_row_index]
                        currency_value = str(first_data_row.iloc[currency_col_index]).strip().upper()
                        
                        logger.info(
                            f"[CURRENCY DEBUG] Currency column found at index {currency_col_index}, value: {currency_value}"
                        )
                        
                        # Match against allowed currencies (code and aliases)
                        if allowed_currencies:
                            if currency_value in allowed_currencies:
                                logger.info(
                                    f"Currency detected from column: {currency_value}",
                                    file=file_path_obj.name
                                )
                                return currency_value, detected_headers
                            # Check aliases using scoped detection
                            detected = currency_detector.detect_currency_from_text_scoped(
                                currency_value, allowed_currencies
                            )
                            if detected:
                                logger.info(
                                    f"Currency detected from column via alias: {detected}",
                                    file=file_path_obj.name
                                )
                                return detected, detected_headers
                
                # If header detection succeeded but no currency in header or column,
                # check first data row price cell
                price_col_index = detected_headers.column_indices.get('price')
                logger.info(
                    f"[CURRENCY DEBUG] Price column index: {price_col_index}",
                    header_row_index=detected_headers.header_row_index,
                    all_columns=detected_headers.column_indices
                )
                
                if price_col_index is not None:
                    # Read first data row (row after header)
                    data_row_index = detected_headers.header_row_index + 1
                    
                    logger.info(
                        f"[CURRENCY DEBUG] Attempting to read first data row at index: {data_row_index}",
                        df_peek_length=len(df_peek)
                    )
                    
                    if data_row_index < len(df_peek):
                        first_data_row = df_peek.iloc[data_row_index]
                        price_value = first_data_row.iloc[price_col_index]
                        
                        logger.info(
                            f"[CURRENCY DEBUG] Raw price value extracted",
                            price_value=repr(price_value),
                            value_type=type(price_value).__name__,
                            is_na=pd.notna(price_value)
                        )
                        
                        # NEW: For Excel files, check cell number format
                        if file_ext in ['.xlsx', '.xls']:
                            try:
                                from openpyxl import load_workbook
                                
                                logger.info("[CURRENCY DEBUG] Opening Excel file to read cell formats")
                                wb = load_workbook(filename=file_path, data_only=False, read_only=True)
                                ws = wb.active
                                
                                if ws:
                                    # Get the actual cell (1-indexed in openpyxl)
                                    cell_row = data_row_index + 1  # Convert to 1-indexed
                                    cell_col = price_col_index + 1  # Convert to 1-indexed
                                    
                                    # Access cell and get its number format
                                    cell = ws.cell(row=cell_row, column=cell_col)
                                    number_format = cell.number_format
                                    
                                    # Get column letter safely (handles merged cells)
                                    from openpyxl.utils import get_column_letter
                                    col_letter = get_column_letter(cell_col)
                                    
                                    logger.info(
                                        f"[CURRENCY DEBUG] Cell number format",
                                        cell_ref=f"{col_letter}{cell_row}",
                                        number_format=number_format
                                    )
                                    
                                    if number_format and number_format != 'General':
                                        # Extract currency from format string
                                        currency_from_format = self._extract_currency_from_format(number_format)
                                        
                                        if currency_from_format:
                                            logger.info(
                                                f"[CURRENCY DEBUG] Currency extracted from format: {repr(currency_from_format)}"
                                            )
                                            
                                            # Use currency detector to identify the currency
                                            if allowed_currencies:
                                                # Scoped detection - check all currencies with this symbol
                                                detected = currency_detector.detect_currency_from_symbol_scoped(
                                                    currency_from_format, 
                                                    allowed_currencies
                                                )
                                                if detected:
                                                    logger.info(
                                                        f"Currency detected from cell format: {detected}",
                                                        file=file_path_obj.name,
                                                        format=number_format
                                                    )
                                                    wb.close()
                                                    return detected, detected_headers
                                            else:
                                                # Unrestricted detection - use standard method
                                                detected = currency_detector.detect_currency_from_symbol(currency_from_format)
                                                if detected:
                                                    logger.info(
                                                        f"Currency detected from cell format: {detected}",
                                                        file=file_path_obj.name,
                                                        format=number_format
                                                    )
                                                    wb.close()
                                                    return detected, detected_headers
                                
                                wb.close()
                            except Exception as e:
                                logger.debug(f"[CURRENCY DEBUG] Failed to read cell format: {e}")
                                # Continue to value-based detection
                        elif file_ext in ['.xlsb', '.xls']:
                            logger.debug(f"[CURRENCY DEBUG] Skipping format detection for {file_ext} file (not supported by openpyxl)")
                        
                        # Continue with existing value-based detection
                        if pd.notna(price_value):
                            price_str = str(price_value).strip()
                            logger.info(
                                f"[CURRENCY DEBUG] Price string converted",
                                price_str=repr(price_str),
                                length=len(price_str),
                                has_dollar=('$' in price_str),
                                first_50_chars=price_str[:50] if len(price_str) > 50 else price_str
                            )
                            
                            if price_str:
                                # Try to detect currency from symbol
                                logger.info(
                                    f"[CURRENCY DEBUG] Calling detect_currency_from_symbol with: {repr(price_str[:100])}"
                                )
                                currency_from_symbol = currency_detector.detect_currency_from_symbol(price_str)
                                
                                if currency_from_symbol:
                                    logger.info(
                                        f"[CURRENCY DEBUG] ✓ Symbol detection SUCCESS: {currency_from_symbol}",
                                        file=file_path_obj.name,
                                        price_value=price_str[:50]
                                    )
                                    return currency_from_symbol, detected_headers
                                else:
                                    logger.info(
                                        f"[CURRENCY DEBUG] ✗ Symbol detection FAILED - no matching symbol found",
                                        price_str=repr(price_str[:100])
                                    )
                                    # Fallback: check if value is a currency code (sub-header row pattern)
                                    if allowed_currencies:
                                        currency_from_text = currency_detector.detect_currency_from_text_scoped(
                                            price_str, allowed_currencies
                                        )
                                        if currency_from_text:
                                            logger.info(
                                                f"Currency detected from sub-header row: {currency_from_text}",
                                                file=file_path_obj.name,
                                                cell_value=price_str
                                            )
                                            return currency_from_text, detected_headers
                            else:
                                logger.info("[CURRENCY DEBUG] Price string is empty after strip")
                        else:
                            logger.info("[CURRENCY DEBUG] Price value is NA/null")
                    else:
                        logger.info(
                            f"[CURRENCY DEBUG] Cannot read data row - index {data_row_index} >= df length {len(df_peek)}"
                        )
                else:
                    logger.info("[CURRENCY DEBUG] Price column not found in detected headers")
                
            except ParsingError as e:
                logger.debug(f"Header detection failed during currency peek: {e}")
                # Continue - we'll return None, but headers may still be useful
                detected_headers = None
            
            logger.info(f"No currency detected in file peek: {file_path_obj.name}")
            return None, detected_headers
            
        except Exception as e:
            logger.warning(f"Error peeking file for currency: {e}")
            return None, None
    
    def _is_json_file(self, file_path: str) -> bool:
        """
        Check if file contains JSON content (even if extension is .xlsx).
        
        API scrapers save JSON with .xlsx extension for duplicate detection consistency.
        This method peeks at the file content to determine actual format.
        
        Args:
            file_path: Path to file to check
            
        Returns:
            True if file contains JSON, False otherwise
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                # Read first few characters to detect JSON
                first_chars = f.read(10).strip()
                # JSON files start with [ or {
                return first_chars.startswith('[') or first_chars.startswith('{')
        except (UnicodeDecodeError, IOError):
            # If we can't read as text, it's likely binary (Excel)
            return False
    
    def stream_parse_to_csv(
        self,
        file_path: str,
        brand_config: Dict,
        csv_writer,
        transform_func,
        matched_brand_text: Optional[str] = None,
        currency_detector: Optional[Any] = None,
        detected_headers: Optional[DetectedHeaders] = None
    ) -> tuple[int, int, List[str]]:
        """
        Parse Excel/CSV/JSON and stream directly to CSV writer (no memory accumulation).
        
        Detects actual file format by content, not just extension (supports JSON saved as .xlsx).
        Uses intelligent header detection to identify columns dynamically.
        
        Args:
            file_path: Path to input file
            brand_config: Brand configuration (columns no longer required)
            csv_writer: CSV writer object (already opened)
            transform_func: Function to transform each item before writing
            matched_brand_text: Brand text matched in email (for <BRAND> substitution)
            currency_detector: CurrencyDetector instance (for <CURRENCY_CODE> substitution)
            detected_headers: Optional pre-detected headers to skip redundant header parsing
            
        Returns:
            Tuple of (total_rows, valid_rows, errors_list)
        """
        file_ext = Path(file_path).suffix.lower()
        
        # Detect headers for Excel/CSV files (JSON uses field names directly)
        column_map: Dict[str, int] = {}
        header_row_index = 0
        warnings: List[str] = []
        
        if file_ext in ['.xlsx', '.xls', '.xlsb', '.csv', '.txt']:
            # Only detect headers for Excel/CSV/TXT (not JSON)
            if not (file_ext in ['.xlsx', '.xls', '.xlsb'] and self._is_json_file(file_path)):
                # Use cached headers if provided and valid, otherwise detect
                if detected_headers and detected_headers.is_valid():
                    logger.info(
                        f"Using cached headers: {len(detected_headers.column_indices)} fields at row {detected_headers.header_row_index}"
                    )
                    column_map = detected_headers.column_indices
                    header_row_index = detected_headers.header_row_index
                    warnings = detected_headers.warnings
                else:
                    # Use intelligent header detection (pass file path, detector reads it)
                    detected_headers = self.header_detector.detect_headers(
                        file_path, 
                        matched_brand_text, 
                        currency_detector
                    )
                    
                    if not detected_headers.is_valid():
                        missing = ', '.join(detected_headers.missing_required)
                        raise ParsingError(f"Missing required columns: {missing}")
                    
                    column_map = detected_headers.column_indices
                    header_row_index = detected_headers.header_row_index
                    warnings = detected_headers.warnings
                    
                    logger.info(
                        f"Header detection successful: {len(column_map)} fields detected at row {header_row_index}"
                    )
        
        # For .xlsx files, check if it's actually JSON content (API scrapers save JSON as .xlsx)
        if file_ext in ['.xlsx', '.xls', '.xlsb']:
            if self._is_json_file(file_path):
                logger.info(f"Detected JSON content in Excel file, using JSON parser")
                return self._stream_json_to_csv(file_path, brand_config, csv_writer, transform_func)
            else:
                return self._stream_excel_to_csv(
                    file_path, brand_config, csv_writer, transform_func,
                    column_map, header_row_index, warnings
                )
        elif file_ext == '.csv':
            return self._stream_csv_to_csv(
                file_path, brand_config, csv_writer, transform_func,
                column_map, header_row_index, warnings
            )
        elif file_ext == '.json':
            return self._stream_json_to_csv(file_path, brand_config, csv_writer, transform_func)
        elif file_ext == '.txt':
            # Treat .txt files as CSV (with delimiter detection)
            return self._stream_csv_to_csv(
                file_path, brand_config, csv_writer, transform_func,
                column_map, header_row_index, warnings
            )
        else:
            raise ParsingError(f"Unsupported file type: {file_ext}")
    
    def parse_file(
        self,
        file_path: str,
        supplier_config: Dict,
        brand_config: Dict,
        matched_brand_text: Optional[str] = None
    ) -> ParsedPriceList:
        """
        Parse a price list file.
        
        Uses intelligent header detection to identify columns dynamically.
        
        Args:
            file_path: Path to Excel or CSV file
            supplier_config: Supplier configuration
            brand_config: Brand-specific configuration
            matched_brand_text: Brand text matched in email (for <BRAND> substitution)
            
        Returns:
            ParsedPriceList with extracted data
            
        Raises:
            ParsingError: If file cannot be parsed
        """
        try:
            logger.info(
                f"Parsing price list file",
                file_path=file_path,
                supplier=supplier_config.get('supplier'),
                brand=brand_config.get('brand')
            )
            
            # Read file based on extension
            file_ext = Path(file_path).suffix.lower()
            
            # Detect headers using intelligent header detection
            column_indices: Dict[str, int] = {}
            header_row_index = 0
            
            if file_ext in ['.xlsx', '.xls', '.xlsb', '.csv', '.txt']:
                # Check if it's actually JSON content (API scrapers save JSON as .xlsx)
                if file_ext in ['.xlsx', '.xls', '.xlsb'] and self._is_json_file(file_path):
                    # JSON files don't need header detection
                    pass
                else:
                    # Use intelligent header detection
                    detected_headers = self.header_detector.detect_headers(
                        file_path,
                        matched_brand_text,
                        None  # currency_detector not available in this context
                    )
                    
                    if not detected_headers.is_valid():
                        missing = ', '.join(detected_headers.missing_required)
                        raise ParsingError(f"Missing required columns: {missing}")
                    
                    column_indices = detected_headers.column_indices
                    header_row_index = detected_headers.header_row_index
                    
                    logger.info(
                        f"Header detection successful: {len(column_indices)} fields detected at row {header_row_index}"
                    )
            
            if file_ext in ['.csv', '.txt']:
                # Use streaming for CSV and TXT files (delimiter auto-detected)
                items, errors, total_rows = self._parse_csv_streaming(
                    file_path, brand_config, column_indices, header_row_index
                )
            elif file_ext in ['.xlsx', '.xls', '.xlsb']:
                if self._is_json_file(file_path):
                    # Handle JSON content in Excel file
                    logger.info(f"Detected JSON content in Excel file")
                    # For JSON, we need to use the JSON parser
                    # This is a fallback - ideally JSON should be handled separately
                    raise ParsingError("JSON content in Excel file not supported in parse_file - use stream_parse_to_csv instead")
                else:
                    # Use chunked reading for Excel files
                    items, errors, total_rows = self._parse_excel_chunked(
                        file_path, brand_config, column_indices, header_row_index
                    )
            else:
                raise ParsingError(f"Unsupported file type: {file_ext}")
            
            logger.info(f"File parsed: {total_rows} total rows")
            
            # Validate required config values
            supplier_name = supplier_config.get('supplier')
            if not supplier_name:
                raise ParsingError("Missing 'supplier' in supplier config")
            
            brand_name = brand_config.get('brand')
            if not brand_name:
                raise ParsingError("Missing 'brand' in brand config")
            
            result = ParsedPriceList(
                items=items,
                supplier=supplier_name,
                brand=brand_name,
                location=brand_config.get('location', 'Unknown'),
                currency=brand_config.get('currency', 'Unknown'),
                filename=Path(file_path).name,
                total_rows=total_rows,
                valid_rows=len(items),
                errors=errors
            )
            
            logger.info(
                f"Parsing complete: {result.valid_rows}/{result.total_rows} valid items",
                supplier=result.supplier,
                brand=result.brand,
                valid_rows=result.valid_rows,
                total_rows=result.total_rows,
                errors=len(errors)
            )
            
            return result
            
        except Exception as e:
            error_msg = f"Failed to parse {file_path}: {str(e)}"
            logger.error(error_msg, error=str(e))
            raise ParsingError(error_msg)
    
    def _parse_items(
        self,
        df: pd.DataFrame,
        brand_config: Dict,
        column_indices: Dict[str, int]
    ) -> tuple[List[PriceListItem], List[str]]:
        """
        Parse items from dataframe using column indices from header detection.
        
        Args:
            df: Pandas DataFrame (header rows already skipped)
            brand_config: Brand configuration
            column_indices: Column indices from header detection (0-based)
            
        Returns:
            Tuple of (items list, errors list)
        """
        items = []
        errors = []
        
        brand = brand_config.get('brand', 'Unknown')
        location = brand_config.get('location', 'Unknown')
        currency = brand_config.get('currency', 'Unknown')
        
        # Get column indices from header detection (already 0-based)
        part_number_col = column_indices.get('partNumber')
        description_col = column_indices.get('description')
        price_col = column_indices.get('price')
        former_part_col = column_indices.get('formerPartNumber')
        supersede_col = column_indices.get('supersedePartNumber')
        
        # Required columns must exist
        if part_number_col is None or description_col is None or price_col is None:
            raise ParsingError("Missing required column mappings: partNumber, description, or price")
        
        # Parse each row
        for idx, row in df.iterrows():
            try:
                item, error = self._create_validated_item(
                    part_number_raw=row.iloc[part_number_col],
                    price_raw=row.iloc[price_col],
                    description_raw=row.iloc[description_col] if description_col is not None else None,
                    former_part_raw=row.iloc[former_part_col] if former_part_col is not None else None,
                    supersede_raw=row.iloc[supersede_col] if supersede_col is not None else None,
                    brand=brand,
                    location=location,
                    currency=currency,
                    row_number=cast(int, idx) + 1,
                    decimal_format=brand_config.get('decimalFormat', 'decimal')
                )
                if error:
                    errors.append(error)
                if item:
                    items.append(item)
                
            except Exception as e:
                errors.append(f"Row {cast(int, idx) + 1}: {str(e)}")
                continue
        
        return items, errors

    def _create_validated_item(
        self,
        part_number_raw: Any,
        price_raw: Any,
        description_raw: Any,
        former_part_raw: Any,
        supersede_raw: Any,
        brand: str,
        location: str,
        currency: str,
        row_number: int,
        decimal_format: str = 'decimal'
    ) -> tuple[Optional[PriceListItem], Optional[str]]:
        """
        Centralized validation and PriceListItem creation.
        
        Handles validation for all row types (pandas Series, list/tuple, dict values).
        Implements supersede duplicate check: if supersedePartNumber equals partNumber
        (case-insensitive), supersede is set to None.
        
        Args:
            part_number_raw: Raw part number value (any type)
            price_raw: Raw price value (any type)
            description_raw: Raw description value (any type, optional)
            former_part_raw: Raw former part number value (any type, optional)
            supersede_raw: Raw supersede part number value (any type, optional)
            brand: Brand name
            location: Location string
            currency: Currency code
            row_number: Row number for error reporting
            decimal_format: Price decimal format ('decimal' or 'comma')
            
        Returns:
            (PriceListItem, None) on success
            (None, error_message) on validation failure with reportable error
            (None, None) on skip (missing required fields, silent)
        """
        # Validate part number (required)
        if part_number_raw is None or (isinstance(part_number_raw, float) and pd.isna(part_number_raw)):
            return None, None  # Skip silently
        part_number = str(part_number_raw).strip()
        if not part_number:
            return None, None
        
        # Parse price - use 0.0 as fallback for missing/invalid prices
        price = 0.0
        price_error = None
        if price_raw is not None and not (isinstance(price_raw, float) and pd.isna(price_raw)):
            price_str = str(price_raw).strip()
            if price_str:
                parsed_price = self._parse_price(price_str, decimal_format)
                if parsed_price is not None:
                    price = parsed_price
                else:
                    price_error = f"Row {row_number}: Invalid price format: {price_raw}, using 0.0"
        
        # Extract description (optional)
        description = ""
        if description_raw is not None and not (isinstance(description_raw, float) and pd.isna(description_raw)):
            description = str(description_raw).strip()
        
        # Extract former part number (optional)
        former_part = None
        if former_part_raw is not None and not (isinstance(former_part_raw, float) and pd.isna(former_part_raw)):
            former_part = str(former_part_raw).strip() or None
        
        # Extract and validate supersede part number (optional)
        # Set to None if supersede equals part_number (case-insensitive)
        supersede_part = None
        if supersede_raw is not None and not (isinstance(supersede_raw, float) and pd.isna(supersede_raw)):
            cleaned = str(supersede_raw).strip()
            if cleaned and cleaned.lower() != part_number.lower():
                supersede_part = cleaned
        
        return PriceListItem(
            part_number=part_number,
            description=description,
            price=price,
            former_part_number=former_part,
            supersede_part_number=supersede_part,
            brand=brand,
            location=location,
            currency=currency,
            row_number=row_number
        ), price_error

    def _clean_excel_csv_value(self, value: str) -> str:
        """
        Clean Excel-formatted CSV values.
        
        Excel CSV export often wraps values in ="value" to force text formatting.
        This method strips those markers: ="36122461831" -> 36122461831
        
        Args:
            value: Raw CSV cell value
            
        Returns:
            Cleaned value
        """
        value = value.strip()
        
        # Handle Excel text format: ="value" -> value
        if value.startswith('="') and value.endswith('"'):
            value = value[2:-1]  # Remove =" prefix and " suffix
        elif value.startswith('=') and value.startswith('="') == False:
            # Handle malformed cases
            value = value[1:]  # Remove = prefix
        
        # Strip any remaining quotes
        value = value.strip('"').strip("'")
        
        return value.strip()
    
    def _is_row_blank(self, row: Any) -> bool:
        """
        Check if a row is blank (all cells empty or None).
        
        Works with different row types:
        - tuple/list from openpyxl iter_rows
        - list from CSV reader
        - pandas Series
        
        Args:
            row: Row data in any supported format
            
        Returns:
            True if row is blank, False otherwise
        """
        if row is None:
            return True
        
        # Handle pandas Series
        if isinstance(row, pd.Series):
            return row.isna().all() or all(
                str(cell).strip() == '' for cell in row if not pd.isna(cell)
            )
        
        # Handle tuple/list (openpyxl, CSV)
        if isinstance(row, (list, tuple)):
            if len(row) == 0:
                return True
            return all(
                cell is None or (isinstance(cell, float) and pd.isna(cell)) or str(cell).strip() == ''
                for cell in row
            )
        
        return False
    
    def _parse_price(self, price_str: str, decimal_format: str) -> Optional[float]:
        """
        Parse price string to float.
        
        Args:
            price_str: Price string
            decimal_format: Format type ('decimal'/'Decimal', 'comma'/'Comma', etc.)
            
        Returns:
            Float price or None if invalid
        """
        try:
            # Clean Excel CSV formatting first
            price_str = self._clean_excel_csv_value(price_str)
            
            # Remove currency prefixes using config-driven list
            # Prefixes are sorted longest-first to avoid partial matches
            for prefix in self.price_prefixes:
                if price_str.startswith(prefix):
                    price_str = price_str[len(prefix):]
                    break  # Only strip one prefix
            
            price_str = price_str.strip()
            
            # Handle different decimal formats (case-insensitive)
            if decimal_format.lower() == 'comma':
                # European format: 1.234,56 -> 1234.56
                price_str = price_str.replace('.', '').replace(',', '.')
            else:
                # US format: 1,234.56 -> 1234.56
                price_str = price_str.replace(',', '')
            
            price = float(price_str)
            
            # Validate price is positive
            if price < 0:
                return None
            
            return price
            
        except (ValueError, AttributeError):
            return None
    
    def _parse_csv_streaming(
        self,
        file_path: str,
        brand_config: Dict,
        column_indices: Dict[str, int],
        header_row_index: int
    ) -> tuple[List[PriceListItem], List[str], int]:
        """
        Parse CSV file using streaming (row-by-row) to minimize memory usage.
        
        Uses column indices from intelligent header detection.
        
        Args:
            file_path: Path to CSV file
            brand_config: Brand configuration
            column_indices: Column indices from header detection (0-based)
            header_row_index: Row index where header was found (0-based)
            
        Returns:
            Tuple of (items list, errors list, total_rows)
        """
        items: List[PriceListItem] = []
        errors = []
        total_rows = 0
        
        brand = brand_config.get('brand', 'Unknown')
        location = brand_config.get('location', 'Unknown')
        currency = brand_config.get('currency', 'Unknown')
        
        # Get column indices from header detection (already 0-based)
        part_number_col = column_indices.get('partNumber')
        description_col = column_indices.get('description')
        price_col = column_indices.get('price')
        former_part_col = column_indices.get('formerPartNumber')
        supersede_col = column_indices.get('supersedePartNumber')
        
        # Required columns must exist
        if part_number_col is None or description_col is None or price_col is None:
            raise ParsingError("Missing required column mappings: partNumber, description, or price")
        
        logger.info("Starting streaming CSV parse")
        
        try:
            # Detect delimiter by sniffing the first few lines (handles .txt files with various delimiters)
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                sample = f.read(8192)  # Read first 8KB for sniffing
            
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                delimiter = dialect.delimiter
                logger.info(f"Detected delimiter: {repr(delimiter)}")
            except csv.Error:
                # Fall back to comma if sniffing fails
                delimiter = ','
                logger.info("Could not detect delimiter, using comma")
            
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f, delimiter=delimiter, quoting=csv.QUOTE_NONE)
                
                # Skip header row(s) - start from row after header
                for _ in range(header_row_index + 1):
                    next(reader, None)
                
                for row_num, row in enumerate(reader, start=header_row_index + 2):
                    total_rows += 1
                    
                    # Log progress (every 50k rows for cloud efficiency)
                    if total_rows % 50000 == 0:
                        logger.info(
                            f"Processed {total_rows} rows, {len(items)} valid items",
                            rows_processed=total_rows,
                            valid_items=len(items)
                        )
                    
                    try:
                        # Skip if row doesn't have enough columns
                        if len(row) <= max(part_number_col, description_col, price_col):
                            continue
                        
                        # Pre-clean Excel CSV values before validation
                        part_number_raw = self._clean_excel_csv_value(row[part_number_col]) if part_number_col < len(row) else None
                        price_raw = self._clean_excel_csv_value(row[price_col]) if price_col < len(row) else None
                        description_raw = self._clean_excel_csv_value(row[description_col]) if description_col < len(row) else None
                        former_part_raw = self._clean_excel_csv_value(row[former_part_col]) if former_part_col is not None and former_part_col < len(row) else None
                        supersede_raw = self._clean_excel_csv_value(row[supersede_col]) if supersede_col is not None and supersede_col < len(row) else None
                        
                        item, error = self._create_validated_item(
                            part_number_raw=part_number_raw,
                            price_raw=price_raw,
                            description_raw=description_raw,
                            former_part_raw=former_part_raw,
                            supersede_raw=supersede_raw,
                            brand=brand,
                            location=location,
                            currency=currency,
                            row_number=row_num,
                            decimal_format=brand_config.get('decimalFormat', 'decimal')
                        )
                        if error:
                            errors.append(error)
                        if item:
                            items.append(item)
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
        
        except Exception as e:
            raise ParsingError(f"Failed to read CSV file: {str(e)}")
        
        logger.info(
            f"Streaming parse complete: {len(items)}/{total_rows} valid items",
            valid_items=len(items),
            total_rows=total_rows,
            errors=len(errors)
        )
        
        return items, errors, total_rows
    
    def _parse_excel_chunked(
        self,
        file_path: str,
        brand_config: Dict,
        column_indices: Dict[str, int],
        header_row_index: int
    ) -> tuple[List[PriceListItem], List[str], int]:
        """
        Parse Excel file with optimized memory usage.
        
        Uses column indices from intelligent header detection.
        
        Note: pandas read_excel() doesn't support chunksize parameter,
        so we use read-only mode and iterator approach with openpyxl.
        
        Args:
            file_path: Path to Excel file
            brand_config: Brand configuration
            column_indices: Column indices from header detection (0-based)
            header_row_index: Row index where header was found (0-based)
            
        Returns:
            Tuple of (items list, errors list, total_rows)
        """
        logger.info("Starting Excel parse with memory optimization")
        
        # Check file extension to determine parsing method
        file_ext = Path(file_path).suffix.lower()
        
        # .xlsb and .xls files require special engines via pandas (openpyxl doesn't support them)
        if file_ext == '.xlsb':
            logger.info("Detected .xlsb file, using pandas with pyxlsb engine")
            return self._parse_excel_with_pandas(file_path, brand_config, column_indices, header_row_index, engine='pyxlsb')
        elif file_ext == '.xls':
            logger.info("Detected .xls file, using pandas with xlrd engine")
            return self._parse_excel_with_pandas(file_path, brand_config, column_indices, header_row_index, engine='xlrd')
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook in read-only mode (memory efficient)
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            
            if ws is None:
                raise ParsingError("Workbook has no active worksheet")
            
            items: List[PriceListItem] = []
            errors = []
            total_rows = 0
            
            brand = brand_config.get('brand', 'Unknown')
            location = brand_config.get('location', 'Unknown')
            currency = brand_config.get('currency', 'Unknown')
            
            # Get column indices from header detection (already 0-based)
            part_number_col = column_indices.get('partNumber')
            description_col = column_indices.get('description')
            price_col = column_indices.get('price')
            former_part_col = column_indices.get('formerPartNumber')
            supersede_col = column_indices.get('supersedePartNumber')
            
            # Required columns must exist
            if part_number_col is None or description_col is None or price_col is None:
                raise ParsingError("Missing required column mappings: partNumber, description, or price")
            
            logger.info("Starting row-by-row Excel parse")
            
            # Iterate rows (read-only mode streams from disk)
            # Skip header row(s) - start from row after header
            row_iterator = ws.iter_rows(values_only=True)
            for _ in range(header_row_index + 1):
                next(row_iterator, None)
            
            for row_num, row in enumerate(row_iterator, start=header_row_index + 2):
                total_rows += 1
                
                # Log progress (every 50k rows for cloud efficiency)
                if total_rows % 50000 == 0:
                    logger.info(
                        f"Processed {total_rows} rows, {len(items)} valid items",
                        rows_processed=total_rows,
                        valid_items=len(items)
                    )
                
                try:
                    # Skip if row doesn't have enough columns
                    if not row or len(row) <= max(part_number_col, description_col, price_col):
                        continue
                    
                    item, error = self._create_validated_item(
                        part_number_raw=row[part_number_col] if part_number_col < len(row) else None,
                        price_raw=row[price_col] if price_col < len(row) else None,
                        description_raw=row[description_col] if description_col < len(row) else None,
                        former_part_raw=row[former_part_col] if former_part_col is not None and former_part_col < len(row) else None,
                        supersede_raw=row[supersede_col] if supersede_col is not None and supersede_col < len(row) else None,
                        brand=brand,
                        location=location,
                        currency=currency,
                        row_number=row_num,
                        decimal_format=brand_config.get('decimalFormat', 'decimal')
                    )
                    if error:
                        errors.append(error)
                    if item:
                        items.append(item)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            wb.close()
            
            logger.info(
                f"Excel parse complete: {len(items)}/{total_rows} valid items",
                valid_items=len(items),
                total_rows=total_rows,
                errors=len(errors)
            )
            
            return items, errors, total_rows
            
        except ImportError:
            # Fallback to pandas if openpyxl not available
            logger.warning("openpyxl not available, falling back to pandas")
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
            # Skip header row(s) and use column indices from header detection
            if header_row_index >= 0:
                df = df.iloc[header_row_index + 1:]
            items, errors = self._parse_items(df, brand_config, column_indices)
            return items, errors, len(df)
        
        except Exception as e:
            logger.error(f"Excel parsing failed: {str(e)}")
            raise ParsingError(f"Failed to parse Excel file: {str(e)}")
    
    def _parse_excel_with_pandas(
        self,
        file_path: str,
        brand_config: Dict,
        column_indices: Dict[str, int],
        header_row_index: int,
        engine: Literal['xlrd', 'openpyxl', 'odf', 'pyxlsb'] = 'openpyxl'
    ) -> tuple[List[PriceListItem], List[str], int]:
        """
        Parse Excel file using pandas (for .xlsb files or as fallback).
        
        Uses column indices from intelligent header detection.
        
        Args:
            file_path: Path to Excel file
            brand_config: Brand configuration
            column_indices: Column indices from header detection (0-based)
            header_row_index: Row index where header was found (0-based)
            engine: Pandas engine to use ('pyxlsb' for .xlsb, 'openpyxl' for .xlsx)
            
        Returns:
            Tuple of (items list, errors list, total_rows)
        """
        logger.info(f"Parsing Excel file with pandas (engine={engine})")
        
        try:
            # Read entire file with pandas
            df = pd.read_excel(file_path, header=None, engine=engine)
            
            # Skip header row(s)
            if header_row_index >= 0:
                df = df.iloc[header_row_index + 1:]
            
            # Use the existing _parse_items method with column indices
            items, errors = self._parse_items(df, brand_config, column_indices)
            
            return items, errors, len(df)
            
        except Exception as e:
            logger.error(f"Pandas Excel parsing failed: {str(e)}")
            raise ParsingError(f"Failed to parse Excel file with pandas: {str(e)}")
    
    def _stream_excel_to_csv(
        self,
        file_path: str,
        brand_config: Dict,
        csv_writer,
        transform_func,
        column_indices: Dict[str, int],
        header_row_index: int,
        warnings: List[str]
    ) -> tuple[int, int, List[str]]:
        """
        Stream Excel rows directly to CSV writer (no memory accumulation).
        
        Args:
            file_path: Path to Excel file
            brand_config: Brand configuration
            csv_writer: CSV writer object
            transform_func: Function to transform PriceListItem to dict
            column_indices: Detected column indices (0-based)
            header_row_index: Row index of header (0-based)
            warnings: Existing warnings from header detection
            
        Returns:
            Tuple of (total_rows, valid_rows, errors_list)
        """
        logger.info("Starting streaming Excel parse")
        
        # Check file extension to determine parsing method
        file_ext = Path(file_path).suffix.lower()
        
        # .xlsb and .xls files require special engines via pandas (openpyxl doesn't support them)
        if file_ext == '.xlsb':
            logger.info("Detected .xlsb file, using pandas-based streaming")
            return self._stream_excel_to_csv_pandas(
                file_path, brand_config, csv_writer, transform_func,
                column_indices, header_row_index, warnings, engine='pyxlsb'
            )
        elif file_ext == '.xls':
            logger.info("Detected .xls file, using pandas-based streaming")
            return self._stream_excel_to_csv_pandas(
                file_path, brand_config, csv_writer, transform_func,
                column_indices, header_row_index, warnings, engine='xlrd'
            )
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook in read-only mode (streams from disk)
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            
            if ws is None:
                raise ParsingError("Workbook has no active worksheet")
            
            errors = list(warnings)  # Start with header detection warnings
            total_rows = 0
            valid_rows = 0
            
            brand = brand_config.get('brand', 'Unknown')
            location = brand_config.get('location', 'Unknown')
            currency = brand_config.get('currency', 'Unknown')
            
            # Get column indices from detected headers
            part_number_col = column_indices.get('partNumber')
            description_col = column_indices.get('description')
            price_col = column_indices.get('price')
            former_part_col = column_indices.get('formerPartNumber')
            supersede_col = column_indices.get('supersedePartNumber')
            
            # Validate required columns exist
            if part_number_col is None:
                raise ParsingError("Required column 'partNumber' not found in file")
            if price_col is None:
                raise ParsingError("Required column 'price' not found in file")
            
            logger.info("Starting row-by-row streaming parse")
            
            # Early validation: Check first non-header row structure
            early_check_row = None
            for check_row in ws.iter_rows(values_only=True, min_row=2, max_row=2):
                early_check_row = check_row
                break
            
            if early_check_row:
                max_col_needed = max(part_number_col, price_col)
                if description_col is not None:
                    max_col_needed = max(max_col_needed, description_col)
                
                if len(early_check_row) <= max_col_needed:
                    error_msg = (
                        f"File structure mismatch! File has {len(early_check_row)} columns, "
                        f"but config needs column {max_col_needed + 1} (0-indexed: {max_col_needed}). "
                        f"First row: {early_check_row[:10] if len(early_check_row) > 10 else early_check_row}. "
                        f"Config: partNumber={part_number_col}, description={description_col}, price={price_col}"
                    )
                    logger.error(error_msg)
                    raise ParsingError(error_msg)
            
            # Stream rows from Excel and write directly to CSV
            # Skip header row(s) - start from row after header
            consecutive_blank_rows = 0
            
            for row_num, row in enumerate(ws.iter_rows(values_only=True, min_row=header_row_index + 2), start=header_row_index + 2):
                total_rows += 1
                
                # Log progress (every 50k rows for cloud efficiency)
                if total_rows % 50000 == 0:
                    logger.info(
                        f"Processed {total_rows} rows, {valid_rows} written to CSV",
                        rows_processed=total_rows,
                        valid_rows=valid_rows
                    )
                
                try:
                    # Check for blank row - track consecutive blanks for early termination
                    if self._is_row_blank(row):
                        consecutive_blank_rows += 1
                        if consecutive_blank_rows >= self.max_consecutive_blank_rows:
                            logger.info(
                                f"Early termination: {consecutive_blank_rows} consecutive blank rows detected at row {row_num}",
                                total_rows=total_rows,
                                valid_rows=valid_rows,
                                consecutive_blank_rows=consecutive_blank_rows
                            )
                            break
                        continue
                    
                    # Reset consecutive blank counter on non-blank row
                    consecutive_blank_rows = 0
                    
                    # Skip if row doesn't have enough columns
                    max_col_needed = max(part_number_col, price_col)
                    if description_col is not None:
                        max_col_needed = max(max_col_needed, description_col)
                    
                    if len(row) <= max_col_needed:
                        continue
                    
                    item, error = self._create_validated_item(
                        part_number_raw=row[part_number_col] if part_number_col < len(row) else None,
                        price_raw=row[price_col] if price_col < len(row) else None,
                        description_raw=row[description_col] if description_col is not None and description_col < len(row) else None,
                        former_part_raw=row[former_part_col] if former_part_col is not None and former_part_col < len(row) else None,
                        supersede_raw=row[supersede_col] if supersede_col is not None and supersede_col < len(row) else None,
                        brand=brand,
                        location=location,
                        currency=currency,
                        row_number=row_num,
                        decimal_format=brand_config.get('decimalFormat', 'decimal')
                    )
                    if error:
                        errors.append(error)
                    if not item:
                        continue
                    
                    # Transform and write immediately (item goes out of scope after this)
                    transformed = transform_func(item)
                    csv_writer.writerow([
                        transformed['Brand'],
                        transformed['Supplier Name'],
                        transformed['Location'],
                        transformed['Currency'],
                        transformed['Part Number'],
                        transformed['Description'],
                        transformed['FORMER PN'],
                        transformed['SUPERSESSION'],
                        transformed['Price']
                    ])
                    
                    valid_rows += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            wb.close()
            
            logger.info(
                f"Streaming parse complete: {valid_rows}/{total_rows} rows written to CSV",
                valid_rows=valid_rows,
                total_rows=total_rows,
                errors=len(errors)
            )
            
            return total_rows, valid_rows, errors
            
        except Exception as e:
            logger.error(f"Streaming Excel parse failed: {str(e)}")
            raise ParsingError(f"Failed to stream parse Excel file: {str(e)}")
    
    def _stream_json_to_csv(
        self,
        file_path: str,
        brand_config: Dict,
        csv_writer,
        transform_func
    ) -> tuple[int, int, List[str]]:
        """
        Stream JSON data directly to CSV writer (optimized fast path for API scrapers).
        
        This method provides significant performance improvements over JSON→Excel→CSV:
        - No pandas DataFrame creation (saves memory)
        - No Excel serialization/deserialization overhead
        - Direct streaming from JSON to CSV
        - Typically 8-10x faster than the Excel conversion path
        
        Args:
            file_path: Path to JSON file
            brand_config: Brand configuration with column mappings
            csv_writer: CSV writer object
            transform_func: Function to transform PriceListItem to dict
            
        Returns:
            Tuple of (total_rows, valid_rows, errors_list)
        """
        logger.info("Starting streaming JSON parse (optimized path)")
        
        try:
            import ijson
            
            errors: List[str] = []
            total_rows = 0
            valid_rows = 0
            
            brand = brand_config.get('brand', 'Unknown')
            location = brand_config.get('location', 'Unknown')
            currency = brand_config.get('currency', 'Unknown')
            
            # Phase 1: Detect JSON structure by peeking at file
            with open(file_path, 'rb') as peek_file:
                # Read first 500 bytes to detect wrapper format
                peek_data = peek_file.read(500).decode('utf-8', errors='ignore')
                has_data_wrapper = '"Data"' in peek_data[:200] or "'Data'" in peek_data[:200]
                
            # Determine ijson path for streaming
            json_path = 'Data.item' if has_data_wrapper else 'item'
            logger.info(f"Detected JSON structure: {'wrapped in Data key' if has_data_wrapper else 'direct array'}")
            
            # Phase 2: Buffer first 10 records to detect format
            buffered_records = []
            with open(file_path, 'rb') as f:
                records_iter = ijson.items(f, json_path)
                for i, record in enumerate(records_iter):
                    buffered_records.append(record)
                    if i >= 9:  # Buffer 10 records (0-9)
                        break
            
            if not buffered_records:
                logger.warning("JSON file contains no data records")
                return 0, 0, []
            
            # Detect record format and handle blank rows
            is_array_of_arrays = isinstance(buffered_records[0], list)
            
            # JSON files use intelligent header detection (no column config needed)
            
            if is_array_of_arrays:
                # Array-of-arrays format: first row is header, subsequent rows are data arrays
                # ALWAYS uses intelligent header detection - column config is IGNORED
                
                # Skip blank rows at the beginning (supplier data quality issue)
                first_non_blank_idx = 0
                for idx, row in enumerate(buffered_records):
                    if isinstance(row, list) and any(cell for cell in row if cell):
                        first_non_blank_idx = idx
                        logger.info(f"Found first non-blank row at index {idx} (skipped {idx} blank rows)")
                        break
                
                header_row = buffered_records[first_non_blank_idx]
                logger.info(f"Detected array-of-arrays format with header: {header_row}")
                # Data rows start after header in buffered records
                buffered_data_rows = buffered_records[first_non_blank_idx + 1:]
                
                # Always use intelligent header detection for JSON array-of-arrays format
                # Column indices from config are NOT used - we detect from header names automatically
                # Convert header row list to pandas Series for header detection
                header_series = pd.Series(header_row)
                
                logger.info("Using intelligent header detection for JSON array-of-arrays format (column config ignored)")
                # Use header detector to map header names to field names
                # Note: currency_detector not available in this method, but not needed for simple headers
                detected_headers, _price_candidates = self.header_detector._try_detect_headers_in_row(
                    row=header_series,
                    row_index=0,
                    matched_brand_text=brand_config.get('brand'),
                    currency_detector=None  # Not needed for simple headers like 'part', 'description', 'price'
                )
                
                if not detected_headers or not detected_headers.is_valid():
                    missing = ', '.join(detected_headers.missing_required) if detected_headers else 'unknown'
                    raise ParsingError(
                        f"Header detection failed for JSON array-of-arrays format. "
                        f"Missing required fields: {missing}. "
                        f"Header row: {header_row}"
                    )
                
                # HeaderDetector returns 0-based indices which match array indices directly
                # No conversion needed - use detected indices as-is
                column_indices = detected_headers.column_indices.copy()
                
                # Note: Optional fields (formerPartNumber, supersedePartNumber) may not be in column_indices
                # We'll use .get() to handle missing fields
                
                logger.info(f"Using array index mapping from header detection: {column_indices}")
                
                # Validate required indices
                if column_indices.get('partNumber') is None or column_indices.get('price') is None:
                    raise ParsingError(f"Missing required column indices from header detection. Detected: {column_indices}")
                
                logger.info(f"Processing array-of-arrays format with streaming")
                
                # Process buffered data rows first
                row_num = first_non_blank_idx + 2  # Start after header
                for record in buffered_data_rows:
                    total_rows += 1
                    
                    # Log progress (every 50k rows for cloud efficiency)
                    if total_rows % 50000 == 0:
                        logger.info(
                            f"Processed {total_rows} rows, {valid_rows} written to CSV",
                            rows_processed=total_rows,
                            valid_rows=valid_rows
                        )
                    
                    try:
                        # Skip if not a list (array)
                        if not isinstance(record, list):
                            errors.append(f"Row {row_num}: Expected array, got {type(record)}")
                            continue
                        
                        # Pre-process description to strip HTML tags if present
                        description_raw = None
                        if column_indices['description'] is not None and column_indices['description'] < len(record):
                            desc_val = record[column_indices['description']]
                            if desc_val:
                                # Strip HTML tags if present (e.g., "<p>WIRING ASSY</p>" -> "WIRING ASSY")
                                description_raw = re.sub(r'<[^>]+>', '', str(desc_val))
                        
                        former_col = column_indices.get('formerPartNumber')
                        supersede_col = column_indices.get('supersedePartNumber')
                        
                        item, error = self._create_validated_item(
                            part_number_raw=record[column_indices['partNumber']] if column_indices['partNumber'] is not None and column_indices['partNumber'] < len(record) else None,
                            price_raw=record[column_indices['price']] if column_indices['price'] is not None and column_indices['price'] < len(record) else None,
                            description_raw=description_raw,
                            former_part_raw=record[former_col] if former_col is not None and former_col < len(record) else None,
                            supersede_raw=record[supersede_col] if supersede_col is not None and supersede_col < len(record) else None,
                            brand=brand,
                            location=location,
                            currency=currency,
                            row_number=row_num,
                            decimal_format=brand_config.get('decimalFormat', 'decimal')
                        )
                        if error:
                            errors.append(error)
                        if not item:
                            continue
                        
                        # Transform and write immediately
                        transformed = transform_func(item)
                        csv_writer.writerow([
                            transformed['Brand'],
                            transformed['Supplier Name'],
                            transformed['Location'],
                            transformed['Currency'],
                            transformed['Part Number'],
                            transformed['Description'],
                            transformed['FORMER PN'],
                            transformed['SUPERSESSION'],
                            transformed['Price']
                        ])
                        
                        valid_rows += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
                    
                    row_num += 1
                
                # Stream remaining records from file (skipping already buffered ones)
                logger.info(f"Streaming remaining records from file (buffered first {len(buffered_records)} records)")
                with open(file_path, 'rb') as f:
                    records_iter = ijson.items(f, json_path)
                    
                    # Skip records we already processed (buffered + header)
                    skip_count = len(buffered_records)
                    for _ in range(skip_count):
                        try:
                            next(records_iter)
                        except StopIteration:
                            break
                    
                    # Process remaining records
                    for record in records_iter:
                        total_rows += 1
                        
                        # Log progress (every 50k rows for cloud efficiency)
                        if total_rows % 50000 == 0:
                            logger.info(
                                f"Processed {total_rows} rows, {valid_rows} written to CSV",
                                rows_processed=total_rows,
                                valid_rows=valid_rows
                            )
                        
                        try:
                            # Skip if not a list (array)
                            if not isinstance(record, list):
                                errors.append(f"Row {row_num}: Expected array, got {type(record)}")
                                row_num += 1
                                continue
                            
                            # Pre-process description to strip HTML tags if present
                            description_raw = None
                            if column_indices['description'] is not None and column_indices['description'] < len(record):
                                desc_val = record[column_indices['description']]
                                if desc_val:
                                    # Strip HTML tags if present (e.g., "<p>WIRING ASSY</p>" -> "WIRING ASSY")
                                    description_raw = re.sub(r'<[^>]+>', '', str(desc_val))
                            
                            former_col = column_indices.get('formerPartNumber')
                            supersede_col = column_indices.get('supersedePartNumber')
                            
                            item, error = self._create_validated_item(
                                part_number_raw=record[column_indices['partNumber']] if column_indices['partNumber'] is not None and column_indices['partNumber'] < len(record) else None,
                                price_raw=record[column_indices['price']] if column_indices['price'] is not None and column_indices['price'] < len(record) else None,
                                description_raw=description_raw,
                                former_part_raw=record[former_col] if former_col is not None and former_col < len(record) else None,
                                supersede_raw=record[supersede_col] if supersede_col is not None and supersede_col < len(record) else None,
                                brand=brand,
                                location=location,
                                currency=currency,
                                row_number=row_num,
                                decimal_format=brand_config.get('decimalFormat', 'decimal')
                            )
                            if error:
                                errors.append(error)
                            if not item:
                                row_num += 1
                                continue
                            
                            # Transform and write immediately
                            transformed = transform_func(item)
                            csv_writer.writerow([
                                transformed['Brand'],
                                transformed['Supplier Name'],
                                transformed['Location'],
                                transformed['Currency'],
                                transformed['Part Number'],
                                transformed['Description'],
                                transformed['FORMER PN'],
                                transformed['SUPERSESSION'],
                                transformed['Price']
                            ])
                            
                            valid_rows += 1
                            
                        except Exception as e:
                            errors.append(f"Row {row_num}: {str(e)}")
                        finally:
                            row_num += 1
                
            else:
                # Object format: rows are dictionaries
                # Always use intelligent auto-detection - column config is IGNORED
                
                # Use first buffered record for field mapping detection
                first_record = buffered_records[0]
                
                # Auto-detect field names from first record
                field_mapping = self._auto_detect_json_fields(first_record)
                logger.info(f"Auto-detected JSON fields: {field_mapping}")
                
                # Validate required fields
                if not all([field_mapping.get('partNumber'), field_mapping.get('price')]):
                    raise ParsingError(f"Missing required field mappings. Found: {field_mapping}")
                
                logger.info(f"Processing JSON records (object format) with streaming")
                
                # Process buffered records first
                row_num = 1
                for record in buffered_records:
                    total_rows += 1
                    
                    # Log progress (every 50k rows for cloud efficiency)
                    if total_rows % 50000 == 0:
                        logger.info(
                            f"Processed {total_rows} rows, {valid_rows} written to CSV",
                            rows_processed=total_rows,
                            valid_rows=valid_rows
                        )
                    
                    try:
                        # Skip if not a dict
                        if not isinstance(record, dict):
                            errors.append(f"Row {row_num}: Expected dict, got {type(record)}")
                            continue
                        
                        # Extract raw values using field mapping
                        item, error = self._create_validated_item(
                            part_number_raw=record.get(field_mapping['partNumber']),
                            price_raw=record.get(field_mapping['price']),
                            description_raw=record.get(field_mapping['description']) if field_mapping.get('description') else None,
                            former_part_raw=record.get(field_mapping['formerPartNumber']) if field_mapping.get('formerPartNumber') else None,
                            supersede_raw=record.get(field_mapping['supersedePartNumber']) if field_mapping.get('supersedePartNumber') else None,
                            brand=brand,
                            location=location,
                            currency=currency,
                            row_number=row_num,
                            decimal_format=brand_config.get('decimalFormat', 'decimal')
                        )
                        if error:
                            errors.append(error)
                        if not item:
                            continue
                        
                        # Transform and write immediately (item goes out of scope after this)
                        transformed = transform_func(item)
                        csv_writer.writerow([
                            transformed['Brand'],
                            transformed['Supplier Name'],
                            transformed['Location'],
                            transformed['Currency'],
                            transformed['Part Number'],
                            transformed['Description'],
                            transformed['FORMER PN'],
                            transformed['SUPERSESSION'],
                            transformed['Price']
                        ])
                        
                        valid_rows += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
                    
                    row_num += 1
                
                # Stream remaining records from file (skipping already buffered ones)
                logger.info(f"Streaming remaining records from file (buffered first {len(buffered_records)} records)")
                with open(file_path, 'rb') as f:
                    records_iter = ijson.items(f, json_path)
                    
                    # Skip records we already processed (buffered)
                    skip_count = len(buffered_records)
                    for _ in range(skip_count):
                        try:
                            next(records_iter)
                        except StopIteration:
                            break
                    
                    # Process remaining records
                    for record in records_iter:
                        total_rows += 1
                        
                        # Log progress (every 50k rows for cloud efficiency)
                        if total_rows % 50000 == 0:
                            logger.info(
                                f"Processed {total_rows} rows, {valid_rows} written to CSV",
                                rows_processed=total_rows,
                                valid_rows=valid_rows
                            )
                        
                        try:
                            # Skip if not a dict
                            if not isinstance(record, dict):
                                errors.append(f"Row {row_num}: Expected dict, got {type(record)}")
                                row_num += 1
                                continue
                            
                            # Extract raw values using field mapping
                            item, error = self._create_validated_item(
                                part_number_raw=record.get(field_mapping['partNumber']),
                                price_raw=record.get(field_mapping['price']),
                                description_raw=record.get(field_mapping['description']) if field_mapping.get('description') else None,
                                former_part_raw=record.get(field_mapping['formerPartNumber']) if field_mapping.get('formerPartNumber') else None,
                                supersede_raw=record.get(field_mapping['supersedePartNumber']) if field_mapping.get('supersedePartNumber') else None,
                                brand=brand,
                                location=location,
                                currency=currency,
                                row_number=row_num,
                                decimal_format=brand_config.get('decimalFormat', 'decimal')
                            )
                            if error:
                                errors.append(error)
                            if not item:
                                row_num += 1
                                continue
                            
                            # Transform and write immediately (item goes out of scope after this)
                            transformed = transform_func(item)
                            csv_writer.writerow([
                                transformed['Brand'],
                                transformed['Supplier Name'],
                                transformed['Location'],
                                transformed['Currency'],
                                transformed['Part Number'],
                                transformed['Description'],
                                transformed['FORMER PN'],
                                transformed['SUPERSESSION'],
                                transformed['Price']
                            ])
                            
                            valid_rows += 1
                            
                        except Exception as e:
                            errors.append(f"Row {row_num}: {str(e)}")
                        finally:
                            row_num += 1
            
            logger.info(
                f"Streaming JSON parse complete: {valid_rows}/{total_rows} rows written to CSV",
                valid_rows=valid_rows,
                total_rows=total_rows,
                errors=len(errors)
            )
            
            return total_rows, valid_rows, errors
            
        except Exception as e:
            logger.error(f"Streaming JSON parse failed: {str(e)}")
            raise ParsingError(f"Failed to stream parse JSON file: {str(e)}")
    
    def _auto_detect_json_fields(self, sample_record: Dict[str, Any]) -> Dict[str, Optional[str]]:
        """
        Auto-detect JSON field names from a sample record using FieldNameDetector.
        
        Uses the same intelligent field detection as HeaderDetector, but for JSON keys.
        All field patterns are loaded from column_mapping_config.json (DRY principle).
        
        Args:
            sample_record: Sample JSON record to analyze
            
        Returns:
            Dictionary mapping column types to JSON field names (optional fields may be None)
        """
        return self.field_name_detector.detect_fields(sample_record)
    
    def _stream_excel_to_csv_pandas(
        self,
        file_path: str,
        brand_config: Dict,
        csv_writer,
        transform_func,
        column_indices: Dict[str, int],
        header_row_index: int,
        warnings: List[str],
        engine: Literal['xlrd', 'openpyxl', 'odf', 'pyxlsb'] = 'pyxlsb'
    ) -> tuple[int, int, List[str]]:
        """
        Stream Excel rows to CSV using pandas (for .xlsb and .xls files).
        
        Note: This is less memory-efficient than openpyxl streaming but necessary
        for .xlsb and .xls files which openpyxl doesn't support.
        
        Args:
            file_path: Path to Excel file (.xlsb or .xls)
            brand_config: Brand configuration
            csv_writer: CSV writer object
            transform_func: Function to transform PriceListItem to dict
            column_indices: Detected column indices (0-based)
            header_row_index: Row index of header (0-based)
            warnings: Existing warnings from header detection
            engine: Pandas engine to use ('pyxlsb' for .xlsb, 'xlrd' for .xls)
            
        Returns:
            Tuple of (total_rows, valid_rows, errors_list)
        """
        logger.info(f"Starting pandas-based Excel streaming (engine={engine})")
        
        try:
            # Read with specified engine
            df = pd.read_excel(file_path, header=None, engine=engine)
            
            errors = list(warnings)
            total_rows = 0
            valid_rows = 0
            
            brand = brand_config.get('brand', 'Unknown')
            location = brand_config.get('location', 'Unknown')
            currency = brand_config.get('currency', 'Unknown')
            
            # Get required column indices
            part_number_col = column_indices.get('partNumber')
            description_col = column_indices.get('description')
            price_col = column_indices.get('price')
            former_part_col = column_indices.get('formerPartNumber')
            supersede_col = column_indices.get('supersedePartNumber')
            
            if part_number_col is None or price_col is None:
                raise ParsingError("Missing required column indices: partNumber and price are required")
            
            # Process rows after header
            consecutive_blank_rows = 0
            
            for idx, row in df.iterrows():
                total_rows += 1
                
                # Skip header and rows before it
                row_idx = int(idx) if isinstance(idx, (int, float)) else 0
                if row_idx <= header_row_index:
                    continue
                
                try:
                    # Check for blank row - track consecutive blanks for early termination
                    if self._is_row_blank(row):
                        consecutive_blank_rows += 1
                        if consecutive_blank_rows >= self.max_consecutive_blank_rows:
                            logger.info(
                                f"Early termination: {consecutive_blank_rows} consecutive blank rows detected at row {row_idx + 1}",
                                total_rows=total_rows,
                                valid_rows=valid_rows,
                                consecutive_blank_rows=consecutive_blank_rows
                            )
                            break
                        continue
                    
                    # Reset consecutive blank counter on non-blank row
                    consecutive_blank_rows = 0
                    
                    item, error = self._create_validated_item(
                        part_number_raw=row.iloc[part_number_col] if part_number_col < len(row) else None,
                        price_raw=row.iloc[price_col] if price_col < len(row) else None,
                        description_raw=row.iloc[description_col] if description_col is not None and description_col < len(row) else None,
                        former_part_raw=row.iloc[former_part_col] if former_part_col is not None and former_part_col < len(row) else None,
                        supersede_raw=row.iloc[supersede_col] if supersede_col is not None and supersede_col < len(row) else None,
                        brand=brand,
                        location=location,
                        currency=currency,
                        row_number=row_idx + 1,
                        decimal_format=brand_config.get('decimalFormat', 'decimal')
                    )
                    if error:
                        errors.append(error)
                    if item:
                        # Transform and write immediately (item goes out of scope after this)
                        transformed = transform_func(item)
                        csv_writer.writerow([
                            transformed['Brand'],
                            transformed['Supplier Name'],
                            transformed['Location'],
                            transformed['Currency'],
                            transformed['Part Number'],
                            transformed['Description'],
                            transformed['FORMER PN'],
                            transformed['SUPERSESSION'],
                            transformed['Price']
                        ])
                        valid_rows += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx + 1}: {str(e)}")
                    continue
            
            logger.info(f"Pandas streaming complete: {valid_rows}/{total_rows} rows valid")
            return total_rows, valid_rows, errors
            
        except Exception as e:
            logger.error(f"Pandas Excel streaming failed: {str(e)}")
            raise ParsingError(f"Failed to stream Excel file with pandas: {str(e)}")
    
    def _stream_csv_to_csv(
        self,
        file_path: str,
        brand_config: Dict,
        csv_writer,
        transform_func,
        column_indices: Dict[str, int],
        header_row_index: int,
        warnings: List[str]
    ) -> tuple[int, int, List[str]]:
        """
        Stream CSV rows directly to output CSV writer.
        
        Similar to _stream_excel_to_csv but for CSV input files.
        Supports CSV and TXT files with auto-delimiter detection.
        """
        logger.info("Starting streaming CSV parse")
        
        try:
            errors: List[str] = list(warnings)  # Start with header detection warnings
            total_rows = 0
            valid_rows = 0
            
            brand = brand_config.get('brand', 'Unknown')
            location = brand_config.get('location', 'Unknown')
            currency = brand_config.get('currency', 'Unknown')
            
            # Get column indices from detected headers
            part_number_col = column_indices.get('partNumber')
            description_col = column_indices.get('description')
            price_col = column_indices.get('price')
            former_part_col = column_indices.get('formerPartNumber')
            supersede_col = column_indices.get('supersedePartNumber')
            
            # Validate required columns exist
            if part_number_col is None:
                raise ParsingError("Required column 'partNumber' not found in file")
            if price_col is None:
                raise ParsingError("Required column 'price' not found in file")
            
            # Detect delimiter by sniffing the file (handles .txt files with various delimiters)
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                sample = f.read(8192)  # Read first 8KB for sniffing
            
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                delimiter = dialect.delimiter
                logger.info(f"Detected delimiter for streaming: {repr(delimiter)}")
            except csv.Error:
                # Fall back to comma if sniffing fails
                delimiter = ','
                logger.info("Could not detect delimiter for streaming, using comma")
            
            # Open and read CSV file with detected delimiter
            with open(file_path, 'r', encoding='utf-8-sig', newline='') as csvfile:
                csv_reader = csv.reader(csvfile, delimiter=delimiter, quoting=csv.QUOTE_NONE)
                
                # Skip to header row and then skip past it
                for _ in range(header_row_index + 1):
                    try:
                        next(csv_reader)
                    except StopIteration:
                        raise ParsingError("CSV file is empty or header row not found")
                
                # Process each data row
                consecutive_blank_rows = 0
                
                for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (after header)
                    total_rows += 1
                    
                    try:
                        # Check for blank row - track consecutive blanks for early termination
                        if self._is_row_blank(row):
                            consecutive_blank_rows += 1
                            if consecutive_blank_rows >= self.max_consecutive_blank_rows:
                                logger.info(
                                    f"Early termination: {consecutive_blank_rows} consecutive blank rows detected at row {row_num}",
                                    total_rows=total_rows,
                                    valid_rows=valid_rows,
                                    consecutive_blank_rows=consecutive_blank_rows
                                )
                                break
                            continue
                        
                        # Reset consecutive blank counter on non-blank row
                        consecutive_blank_rows = 0
                        
                        item, error = self._create_validated_item(
                            part_number_raw=row[part_number_col] if part_number_col < len(row) else None,
                            price_raw=row[price_col] if price_col < len(row) else None,
                            description_raw=row[description_col] if description_col is not None and description_col < len(row) else None,
                            former_part_raw=row[former_part_col] if former_part_col is not None and former_part_col < len(row) else None,
                            supersede_raw=row[supersede_col] if supersede_col is not None and supersede_col < len(row) else None,
                            brand=brand,
                            location=location,
                            currency=currency,
                            row_number=row_num,
                            decimal_format=brand_config.get('decimalFormat', 'decimal')
                        )
                        if error:
                            errors.append(error)
                        if not item:
                            continue
                        
                        # Transform and write immediately (item goes out of scope after this)
                        transformed = transform_func(item)
                        csv_writer.writerow([
                            transformed['Brand'],
                            transformed['Supplier Name'],
                            transformed['Location'],
                            transformed['Currency'],
                            transformed['Part Number'],
                            transformed['Description'],
                            transformed['FORMER PN'],
                            transformed['SUPERSESSION'],
                            transformed['Price']
                        ])
                        
                        valid_rows += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
            
            logger.info(
                f"Streaming CSV parse complete: {valid_rows}/{total_rows} rows written to CSV",
                valid_rows=valid_rows,
                total_rows=total_rows,
                errors=len(errors)
            )
            
            return total_rows, valid_rows, errors
            
        except Exception as e:
            logger.error(f"Streaming CSV parse failed: {str(e)}")
            raise ParsingError(f"Failed to stream parse CSV file: {str(e)}")
